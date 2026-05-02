import pandas as pd
from io import BytesIO

class ExcelService:
    def create_tabel_excel(self, data: list, sheet_name: str = 'Data KOL') -> BytesIO:
        if not data:
            df = pd.DataFrame()
        else:
            df = pd.DataFrame(data)
            
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Access the openpyxl objects for styling
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            
            # Premium Styling
            header_fill = PatternFill(start_color='1A202C', end_color='1A202C', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center', indent=1)
            
            # Border tebal untuk header
            header_border = Border(
                left=Side(style='medium', color='000000'),
                right=Side(style='medium', color='000000'),
                top=Side(style='medium', color='000000'),
                bottom=Side(style='medium', color='000000')
            )

            # Border tipis untuk isi data (kiri, kanan, bawah)
            body_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='D0D0D0'),
                bottom=Side(style='thin', color='000000')
            )

            # Style Headers
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = header_border

            # Hitung jumlah baris dan kolom yang ada data
            max_row = worksheet.max_row
            max_col = worksheet.max_column

            # Adjust column width and style content
            for col_idx, column in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                    
                    # Style body cells
                    if cell.row > 1:
                        # Tentukan border: sel paling kiri/kanan pakai medium, yang lain thin
                        is_first_col = (col_idx == 1)
                        is_last_col = (col_idx == max_col)
                        is_last_row = (cell.row == max_row)

                        left_side = Side(style='medium', color='000000') if is_first_col else Side(style='thin', color='000000')
                        right_side = Side(style='medium', color='000000') if is_last_col else Side(style='thin', color='000000')
                        bottom_side = Side(style='medium', color='000000') if is_last_row else Side(style='thin', color='000000')
                        top_side = Side(style='thin', color='D0D0D0')

                        cell.border = Border(
                            left=left_side,
                            right=right_side,
                            top=top_side,
                            bottom=bottom_side
                        )
                        cell.font = Font(size=10, name='Calibri')
                        
                        # Custom alignment based on content
                        header_val = str(worksheet.cell(row=1, column=col_idx).value).upper()
                        if header_val in ['NAMA', 'USERNAME', 'LINK', 'LINK INSTAGRAM']:
                            cell.alignment = left_alignment
                        else:
                            cell.alignment = center_alignment
                
                adjusted_width = (max_length + 4)
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

        buffer.seek(0)
        return buffer

    def create_tabel_csv(self, data: list) -> BytesIO:
        if not data:
            df = pd.DataFrame()
        else:
            df = pd.DataFrame(data)
            
        buffer = BytesIO()
        # Use utf-8-sig for better compatibility with Excel when opening CSV
        csv_data = df.to_csv(index=False, encoding='utf-8-sig')
        buffer.write(csv_data.encode('utf-8-sig'))
        buffer.seek(0)
        return buffer

excel_service = ExcelService()
